"""
JieDimension Toolkit - 数据导出模块
支持导出Excel报告
Version: 1.0.0
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List
import os
import sys

# 添加项目根目录到路径
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from core.database import Database


class ExcelReportExporter:
    """Excel报告导出器"""
    
    def __init__(self, db: Database):
        """
        初始化导出器
        
        Args:
            db: 数据库实例
        """
        self.db = db
    
    async def export_full_report(
        self,
        output_path: str,
        days: int = 30
    ) -> bool:
        """
        导出完整报告到Excel
        
        Args:
            output_path: 输出文件路径
            days: 统计天数
            
        Returns:
            是否成功
        """
        try:
            # 创建工作簿
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认工作表
            
            # 获取日期范围
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            
            # 1. 概览页
            try:
                await self._create_overview_sheet(wb, start_date, end_date)
            except Exception as e:
                print(f"创建概览页失败: {e}")
                import traceback
                traceback.print_exc()
            
            # 2. AI使用统计页
            try:
                await self._create_ai_stats_sheet(wb, start_date, end_date)
            except Exception as e:
                print(f"创建AI统计页失败: {e}")
                import traceback
                traceback.print_exc()
            
            # 3. 任务统计页
            try:
                await self._create_tasks_sheet(wb, start_date, end_date)
            except Exception as e:
                print(f"创建任务统计页失败: {e}")
                import traceback
                traceback.print_exc()
            
            # 4. 商品列表页
            try:
                await self._create_products_sheet(wb)
            except Exception as e:
                print(f"创建商品列表页失败: {e}")
                import traceback
                traceback.print_exc()
            
            # 保存文件
            wb.save(output_path)
            print(f"✅ 报告已导出到: {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ 导出报告失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    async def _create_overview_sheet(
        self,
        wb: Workbook,
        start_date: datetime,
        end_date: datetime
    ):
        """创建概览页"""
        ws = wb.create_sheet("📊 概览")
        
        # 设置标题
        ws['A1'] = "JieDimension Toolkit - 数据报告"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # 报告信息
        ws['A3'] = "报告日期:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A4'] = "统计周期:"
        ws['B4'] = f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}"
        
        # 获取统计数据
        total_products = await self.db.count_products()
        ai_stats = await self.db.get_ai_stats_summary()
        tasks = await self.db.get_tasks_by_date_range(
            start_date.isoformat(),
            end_date.isoformat()
        )
        
        # 关键指标
        ws['A6'] = "关键指标"
        ws['A6'].font = Font(size=14, bold=True)
        ws.merge_cells('A6:D6')
        
        avg_latency = ai_stats.get('avg_latency') or 0
        success_rate = ai_stats.get('success_rate') or 0
        
        metrics = [
            ("商品总数", total_products),
            ("任务总数", len(tasks)),
            ("AI调用次数", ai_stats.get('total_calls', 0)),
            ("AI成功率", f"{success_rate:.1f}%"),
            ("平均延迟", f"{avg_latency:.2f}秒"),
        ]
        
        row = 7
        for metric_name, value in metrics:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
    
    async def _create_ai_stats_sheet(
        self,
        wb: Workbook,
        start_date: datetime,
        end_date: datetime
    ):
        """创建AI使用统计页"""
        ws = wb.create_sheet("🤖 AI使用统计")
        
        # 标题
        ws['A1'] = "AI使用统计"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # 获取AI调用记录
        ai_calls = await self.db.get_ai_calls(
            start_date.isoformat(),
            end_date.isoformat()
        )
        
        # 按提供商统计
        ws['A3'] = "按提供商统计"
        ws['A3'].font = Font(bold=True)
        
        # 表头
        headers = ['提供商', '调用次数', '成功次数', '成功率', '平均延迟(秒)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 统计数据
        provider_stats = {}
        for call in ai_calls:
            provider = call['provider']
            if provider not in provider_stats:
                provider_stats[provider] = {
                    'total': 0,
                    'success': 0,
                    'latency_sum': 0
                }
            provider_stats[provider]['total'] += 1
            if call['success']:
                provider_stats[provider]['success'] += 1
            provider_stats[provider]['latency_sum'] += call['latency']
        
        # 填充数据
        row = 5
        for provider, stats in provider_stats.items():
            ws[f'A{row}'] = provider
            ws[f'B{row}'] = stats['total']
            ws[f'C{row}'] = stats['success']
            success_rate = (stats['success'] / stats['total'] * 100) if stats['total'] > 0 else 0
            ws[f'D{row}'] = f"{success_rate:.1f}%"
            avg_latency = (stats['latency_sum'] / stats['total']) if stats['total'] > 0 else 0
            ws[f'E{row}'] = f"{avg_latency:.2f}"
            row += 1
        
        # 详细记录
        ws[f'A{row + 2}'] = "详细调用记录"
        ws[f'A{row + 2}'].font = Font(bold=True)
        
        # 表头
        headers = ['时间', '提供商', '模型', '任务类型', '复杂度', '延迟(秒)', '状态']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row + 3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 填充记录（最近50条）
        detail_row = row + 4
        for call in ai_calls[-50:]:
            ws[f'A{detail_row}'] = call.get('created_at', '-')
            ws[f'B{detail_row}'] = call.get('provider', '-')
            ws[f'C{detail_row}'] = call.get('model', '-')
            ws[f'D{detail_row}'] = call.get('task_type', '-')
            ws[f'E{detail_row}'] = call.get('complexity', 0)
            latency = call.get('latency', 0)
            ws[f'F{detail_row}'] = f"{latency:.2f}" if latency is not None else "0.00"
            ws[f'G{detail_row}'] = "✅ 成功" if call.get('success', False) else "❌ 失败"
            detail_row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
    
    async def _create_tasks_sheet(
        self,
        wb: Workbook,
        start_date: datetime,
        end_date: datetime
    ):
        """创建任务统计页"""
        ws = wb.create_sheet("📝 任务统计")
        
        # 标题
        ws['A1'] = "任务统计"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # 获取任务数据
        tasks = await self.db.get_tasks_by_date_range(
            start_date.isoformat(),
            end_date.isoformat()
        )
        
        # 按状态统计
        ws['A3'] = "按状态统计"
        ws['A3'].font = Font(bold=True)
        
        status_count = {}
        for task in tasks:
            status = task['status']
            status_count[status] = status_count.get(status, 0) + 1
        
        row = 4
        for status, count in status_count.items():
            ws[f'A{row}'] = status
            ws[f'B{row}'] = count
            row += 1
        
        # 任务列表
        ws[f'A{row + 2}'] = "任务列表"
        ws[f'A{row + 2}'].font = Font(bold=True)
        
        # 表头
        headers = ['任务名称', '平台', '状态', '进度', '创建时间', '完成时间']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row + 3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 填充数据
        detail_row = row + 4
        for task in tasks:
            ws[f'A{detail_row}'] = task.get('task_name', task.get('type', '-'))
            ws[f'B{detail_row}'] = task.get('platform', '-')
            ws[f'C{detail_row}'] = task.get('status', '-')
            ws[f'D{detail_row}'] = f"{task.get('progress', 0)}%"
            ws[f'E{detail_row}'] = task.get('created_at', '-')
            ws[f'F{detail_row}'] = task.get('completed_at', '-') if task.get('completed_at') else '-'
            detail_row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
    
    async def _create_products_sheet(self, wb: Workbook):
        """创建商品列表页"""
        ws = wb.create_sheet("📦 商品列表")
        
        # 标题
        ws['A1'] = "商品列表"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # 获取商品数据
        products = await self.db.get_products()
        
        # 表头
        headers = ['ID', '标题', '价格', '分类', '平台', '状态', '导入时间']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 填充数据
        row = 4
        for product in products:
            ws[f'A{row}'] = product.get('id', '-')
            ws[f'B{row}'] = product.get('title', '-')
            ws[f'C{row}'] = product.get('price', 0)
            ws[f'D{row}'] = product.get('category', '-') if product.get('category') else '-'
            ws[f'E{row}'] = product.get('platform', '-')
            ws[f'F{row}'] = product.get('status', '-')
            ws[f'G{row}'] = product.get('import_time', '-') if product.get('import_time') else '-'
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
    
    async def export_ai_stats_only(
        self,
        output_path: str,
        days: int = 30
    ) -> bool:
        """
        仅导出AI统计数据
        
        Args:
            output_path: 输出文件路径
            days: 统计天数
            
        Returns:
            是否成功
        """
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            
            # 获取AI调用记录
            ai_calls = await self.db.get_ai_calls(
                start_date.isoformat(),
                end_date.isoformat()
            )
            
            # 转换为DataFrame
            df = pd.DataFrame(ai_calls)
            
            # 导出到Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='AI调用记录', index=False)
            
            print(f"✅ AI统计数据已导出到: {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ 导出AI统计失败: {e}")
            return False


# ===== 测试代码 =====

async def test_export():
    """测试导出功能"""
    print("\n" + "="*60)
    print("🧪 测试Excel报告导出")
    print("="*60)
    
    # 创建数据库连接
    db = Database("data/database.db")
    await db.connect()
    
    # 创建导出器
    exporter = ExcelReportExporter(db)
    
    # 测试1: 导出完整报告
    print("\n🧪 测试1: 导出完整报告")
    output_path = "tests/test_full_report.xlsx"
    success = await exporter.export_full_report(output_path, days=30)
    
    if success and os.path.exists(output_path):
        file_size = os.path.getsize(output_path)
        print(f"✅ 完整报告导出成功 ({file_size} bytes)")
    else:
        print(f"❌ 完整报告导出失败")
    
    # 测试2: 导出AI统计
    print("\n🧪 测试2: 导出AI统计")
    output_path = "tests/test_ai_stats.xlsx"
    success = await exporter.export_ai_stats_only(output_path, days=30)
    
    if success and os.path.exists(output_path):
        file_size = os.path.getsize(output_path)
        print(f"✅ AI统计导出成功 ({file_size} bytes)")
    else:
        print(f"❌ AI统计导出失败")
    
    await db.close()
    print("\n🎉 导出测试完成！")


if __name__ == "__main__":
    import asyncio
    asyncio.run(test_export())

